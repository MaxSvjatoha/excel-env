import os
import sys

import utils.util as utils

import openpyxl

''' File description + Github Copilot context:
This script reads a single template excel file and tries to fill match its sheets to appropriate folders and files in the input folder.
It then reads the data from the input files and saves it to the relevant sheets in the template file. 
'''

# TODO: Code cleanup:
# Make a file retrieval function that reads all the necessary sheets from the template file and returns them as a dict.
# Make a file retrieval function that reads all the necessary sheets from the input files and returns them as a dict.
# Make a function that takes the template sheets and input sheets and matches them to each other (also optional filtering of duplicate matches)
# Make a function that uses the matched sheets to fill in the template file with the data from the input files.

# Settings
cwd = os.getcwd()
parent = os.path.dirname(cwd)

# Get template file
template_folder_name = 'Aktivitetsdata Klimatbokslut.xlsx'
 
# Get input folder list by scanning the parent folder for folders
input_folder = 'Arbetsmapp datainsamling'
input_folder_path = os.path.join(parent, input_folder)
input_folder_names = os.listdir(input_folder_path)

# Remove items that are not folders
input_folder_names = [folder for folder in input_folder_names if os.path.isdir(os.path.join(input_folder_path, folder))]

# Get template folder path
template_folder_path = os.path.join(input_folder_path, template_folder_name)

# Get a list of paths for each folder among the input folder names
input_folder_paths = [os.path.join(input_folder_path, folder) for folder in input_folder_names]

# Go into each input folder path and get a list of all excel files, then save them to a dictionary with the folder name as key
# and the list of excel files as value. If there are no excel files in the folder, skip it.
input_folder_excel_files = {}
input_folder_excel_keys = []
for folder_path in input_folder_paths:
    folder_name = os.path.basename(folder_path)
    folder_files = os.listdir(folder_path)
    folder_excel_files = [file for file in folder_files if file.endswith('.xlsx')]
    if len(folder_excel_files) > 0:
        input_folder_excel_files[folder_name] = folder_excel_files # This will be a subdict with template sheet name as key and itself as the value
        input_folder_excel_keys.append(folder_name)

# Get a list of all sheets in the template file
template_wb = openpyxl.load_workbook(template_folder_path)
template_sheets = template_wb.sheetnames

print(input_folder_excel_keys)
print(utils.normalize_list(data = input_folder_excel_keys))
print(template_sheets)
print(utils.normalize_list(data = template_sheets))

# Match template sheets to input folders - NOTE: normalize_list() is better, but needs a way to trace back the matches to the original, non-normalized list.
#matches = utils.match_lists(utils.normalize_list(data = input_folder_excel_keys), utils.normalize_list(data = template_sheets))
matches = utils.match_lists(input_folder_excel_keys, template_sheets)

# Check "matches" for multiple occurences of the same 'match' key in the subdict. If found, remove the subdict with the lowest 'score' value.
for item in matches.copy(): # Use copy() to avoid RuntimeError: dictionary changed size during iteration
    # Check if items score value is less than 1.0
    if matches[item]['score'] < 1.0:
        # Check if there are multiple occurences of the same 'match' key in the subdict
        if len([match for match in matches if matches[match]['match'] == matches[item]['match']]) > 1:
            # Remove the subdict with the lowest 'score' value
            matches.pop(item)

# We now have a filtered list with associations between template sheets and input folders! Now we want to fill in the subdicts with more relevant data.
# This is now the base dict that we will use to store the data from the input files and then write it to the template file.
base_dict = matches.copy()

print(base_dict)
# Now we want to go through each folder in the input folder and read the excel files in them and store the data inside the base dict.
for folder in input_folder_excel_files:
    # Get the list of excel files in the folder
    excel_files = input_folder_excel_files[folder]
    # Go through each excel file in the list
    for excel_file in excel_files:
        # Get the excel file path
        excel_file_path = os.path.join(input_folder_path, folder, excel_file)
        print(f"Reading file: {excel_file_path}")
        # Read the excel file to a list
        wb = utils.excel_to_workbook(excel_file_path)
        print(wb)
        # Get the sheet names
        sheet_names = wb.sheetnames
        # Go through each sheet in the list and filter out sheets that do not contain 'scope' in the name
        scope_sheets = [sheet for sheet in sheet_names if 'scope' in sheet.lower()]
        # Go through each scope sheet and read the data. If a cell contain an int, store it as a value in a dict, with the precending cell value as a key.
        for scope_sheet in scope_sheets:
            print(f"Reading sheet: {scope_sheet}")
            # Get the sheet
            sheet = wb[scope_sheet]
            # Get the max row and column
            max_row = sheet.max_row
            max_column = sheet.max_column
            # Go through each row and column and read the data
            for row in range(1, max_row + 1):
                for column in range(1, max_column + 1):
                    # Get the cell
                    cell = sheet.cell(row=row, column=column)
                    # Check if the cell contains an int
                    if type(cell.value) == int or type(cell.value) == float:
                        # Get the precending cell
                        precending_cell = sheet.cell(row=row, column=column - 1)
                        # Check if the precending cell contains a string
                        if type(precending_cell.value) == str:
                            # If precending cell contains "specificera", take instead the succeeding cell as the key
                            if "specificera" in precending_cell.value.lower():
                                # Get the succeeding cell
                                succeeding_cell = sheet.cell(row=row, column=column + 1)
                                # Check if the succeeding cell contains a string
                                if type(succeeding_cell.value) == str:
                                    # Store the data in the base dict
                                    base_dict[folder][succeeding_cell.value] = cell.value
                                else:
                                    # The succeeding cell does not contain a string. Skip it.
                                    continue
                            else:
                                # Store the data in the base dict
                                base_dict[folder][precending_cell.value] = cell.value
                        else:
                            # The precending cell does not contain a string. Skip it.
                            continue
                    else:
                        # The cell does not contain an int. Skip it.
                        continue

for item in base_dict:
    print(f"Key: {item}, Value: {base_dict[item]}")

write_dict = base_dict.copy() # dict to write to the template file

#remove 'match' and 'score' keys from the subdicts.
for item in write_dict:
    write_dict[item].pop('match')
    write_dict[item].pop('score')

# Find all unique keys in the subdicts
unique_keys = []
for item in write_dict:
    temp_keys = list(write_dict[item].keys())
    # remove spaces from the keys if it is an the end of the string
    temp_keys = [key.rstrip() for key in temp_keys]
    for key in temp_keys:
        if key not in unique_keys:
            unique_keys.append(key)

print(unique_keys)

# Now, repeat the entire process for the template file, where we want to extract the data for each sheet and store it in a dict using the sheet name as the key.

# Get a list of all sheets in the template file
template_wb = openpyxl.load_workbook(template_folder_path)
template_sheets = template_wb.sheetnames

# Remove 'Sammanställning' from the list
template_sheets.remove('Sammanställning')

# Create a dict to store the data from the template file
template_dict = {}

# Go through each sheet in the template file and read the data
for sheet in template_sheets:
    # Get the sheet
    template_sheet = template_wb[sheet]
    print(f"Reading sheet: {sheet}")
    # Initialize the dict for the sheet in the template dict
    template_dict[sheet] = {}
    # Get the max row and column
    max_row = template_sheet.max_row
    max_column = template_sheet.max_column
    # Go through each row and column and read the data
    for row in range(1, max_row + 1):
        for column in range(1, max_column + 1):
            # Get the cell
            cell = template_sheet.cell(row=row, column=column)
            # Check if the cell contains an int
            if type(cell.value) == int or type(cell.value) == float:
                # Get the precending cell
                precending_cell = template_sheet.cell(row=row, column=column - 1)
                # Check if the precending cell contains a string
                if type(precending_cell.value) == str:
                    # If precending cell contains "specificera", take instead the succeeding cell as the key
                    if "specificera" in precending_cell.value.lower():
                        # Get the succeeding cell
                        succeeding_cell = template_sheet.cell(row=row, column=column + 1)
                        # Check if the succeeding cell contains a string
                        if type(succeeding_cell.value) == str:
                            # Store the data in the template dict
                            template_dict[sheet][succeeding_cell.value] = cell.value
                        else:
                            # The succeeding cell does not contain a string. Skip it.
                            continue
                    else:
                        # Store the data in the template dict
                        template_dict[sheet][precending_cell.value] = cell.value
                else:
                    # The precending cell does not contain a string. Skip it.
                    continue
            else:
                # The cell does not contain an int. Skip it.
                continue

for item in template_dict:
    print(f"Key: {item}, Value: {template_dict[item]}")

# Get unique keys from the template dict
template_unique_keys = []
for item in template_dict:
    temp_keys = list(template_dict[item].keys())
    # remove spaces from the keys if it is an the end of the string
    temp_keys = [key.rstrip() for key in temp_keys]
    for key in temp_keys:
        if key not in template_unique_keys:
            template_unique_keys.append(key)

print(template_unique_keys)
# Now, we have two dicts, one with the data from the excel files and one with the data from the template file.
# We want to compare the data from the excel files with the data from the template file and write the data from the excel files to the template file.

# First, make a copy of the template dict
output_dict = template_dict.copy()
# Go through each sheet in the template file
for sheet in output_dict:
    # Go through each key in the template dict
    for key in output_dict[sheet]:
        # Check if the key is in the base dict
        try:
            if key in write_dict[sheet]:
                print(f"Sheet: {sheet}, Key: {key}")
                print(f"Output dict before: {output_dict[sheet][key]}")
                # If the key is in the base dict, write the value from the base dict to the output dict
                output_dict[sheet][key] = write_dict[sheet][key]
                print(f"Output dict after: {output_dict[sheet][key]}")
        except KeyError:
            # The key is not in the base dict. Skip it.
            continue
        except Exception as e:
            print(f"Error: {e}")
            sys.exit(1)
sys.exit(0)

if __name__ == '__main__':

    print("Script started")

    print("Script finished successfully")
    sys.exit(0)