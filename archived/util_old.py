import os
from openpyxl import Workbook, load_workbook
import numpy as np
import pandas as pd
import re

from typing import List, Dict, Union, Any

import difflib

# NOTE: Internal testing functions are prefixed with an underscore (_) and are not meant for production use
# NOTE 2: Underscore (_) prefix use somewhat inconsistent, should be fixed in a future update

def _generate_excel(cols: int, rows: int, col_names: Union[List, None] = None, save_name: str = 'test.xlsx') -> int:
    '''
    Generate an excel file with 'cols' many columns and 'rows' many rows. 
    Populate it with column headers and random numbers.

    Args:
        cols (int): Number of columns
        rows (int): Number of rows
    Returns:
        int: 0 if successful, 1 if not
    '''

    # Make sure the number of columns and rows are valid
    if cols < 1:
        print('[_generate_excel] Error: Number of columns must be greater than 0')
        return 1
    if rows < 1:
        print('[_generate_excel] Error: Number of rows must be greater than 0')
        return 1
    
    # If column names are provided, make sure they match the number of columns
    if col_names is not None and len(col_names) != cols:
        print('[_generate_excel] Error: Number of column names does not match number of columns')
        return 1
        
    # Make sure the save name is valid
    if save_name == '':
        print('[_generate_excel] Error: Save name cannot be empty')
        return 1
    if not save_name.endswith('.xlsx'):
        save_name += '.xlsx'
        print(f'[_generate_excel] Warning: Save name does not end with .xlsx. Automatically appending .xlsx to save name')

    # Generate the excel file
    try:
        wb = Workbook()
        ws = wb.active

        # Generate column headers
        if col_names is None:
            for i in range(1, cols + 1):
                ws.cell(row=1, column=i).value = 'Column {}'.format(i)
        else:
            for i in range(1, cols + 1):
                ws.cell(row=1, column=i).value = col_names[i - 1]

        # Generate random numbers
        for i in range(2, rows + 2):
            for j in range(1, cols + 1):
                ws.cell(row=i, column=j).value = np.random.randint(1, 100)

        # Save the excel file
        wb.save(os.path.join(os.getcwd(), save_name))
        return 0
    except Exception as e:
        print(e)
        return 1
    
def _get_input_files(path: Union[List[str], str], settings: Dict) -> List[str]:
    '''
    Get all excel file paths in the given path

    Args:
        path (List[str] or str): Path to the folder(s) to search in
        settings (dict): Script settings dictionary

    Returns:
        list: List of excel files
    '''
    if isinstance(path, str):
        path = [path]
    excel_files = []

    for p in path:
        for root, dirs, files in os.walk(p):
            for f in files:
                # If file is an excel file and is part of a subdirectory, add it to the list
                if f.endswith('.xlsx') and not root.split(os.sep)[-1] in settings['Input file folder name']:
                    excel_files.append(os.path.join(root, f))

    return excel_files

# Note: This function is not meant for production use, but this may change in a future update
def _get_input_folders(path: Union[List[str], str], settings: Dict) -> List[str]:
    '''
    Get all folder paths that contain the input excel files

    Args:
        path (List[str] or str): Path to the folder(s) to search in
        settings (dict): Script settings dictionary

    Returns:
        list: List of excel files
    '''
    if isinstance(path, str):
        path = [path]
    excel_folders = []

    for p in path:
        for root, dirs, files in os.walk(p):
            for d in dirs:
                # If dir contains excel files, add it to the list
                if len([f for f in os.listdir(os.path.join(root, d)) if f.endswith('.xlsx')]) > 0:
                    excel_folders.append(os.path.join(root, d))

    return excel_folders

def _get_scope_data(wb: Workbook, sheet: str) -> Dict[str, list]:
    '''
    Get the scope data from the given workbook and sheet

    Args:
        wb (Workbook): Workbook to get the scope data from
        sheet (str): Worksheet to get the scope data from

    Returns:
        dict: TODO
    '''
    result_dict = {}
    sheet = wb[sheet]
    max_row = sheet.max_row
    max_col = sheet.max_column
    
    for row in range(1, max_row + 1):
        # Don't load the first column, since previous cell can't load then
        for col in range(2, max_col + 1):
            prev_cell = sheet.cell(row=row, column=col-1)
            cell = sheet.cell(row=row, column=col)
            next_cell = sheet.cell(row=row, column=col+1)

            # Check if the cell color is FFDDEBF7
            if 'FFDDEBF7'.lower() in cell.fill.start_color.index.lower():
                if next_cell.value is not None:
                    # If it has the same color, then it's the key, else the previous cell is the key
                    if 'FFDDEBF7'.lower() in next_cell.fill.start_color.index.lower():
                        key = next_cell.value
                        #print("[_get_scope_data] next cell is key")
                    else:
                        key = prev_cell.value
                        #print("[_get_scope_data] prev cell is key")
                # If previous cell is not None, then it's a key, else ignore current cell
                elif prev_cell.value is not None:
                    key = prev_cell.value
                else:
                    continue

                # Extra step. Possibly temporary until better matching technique are explored:
                # If key ends with " ", remove it
                if key.endswith(' '):
                    key = key[:-1]
                result_dict[key] = cell.value
    return result_dict

# Will contain several steps, but for now just removes trailing spaces
def preprocess_cell(cell: str) -> str:
    '''
    Preprocess the given cell

    Args:
        cell (str): Cell to preprocess

    Returns:
        str: Preprocessed cell
    '''
    return cell.strip()

def excel_to_list(path: str) -> List[List[str]]:
    '''
    Read an excel file and return a list of lists of strings

    Args:
        path (str): Path to the excel file
    Returns:
        List[List[str]]: List of lists of strings
    '''
    try:
        df = pd.read_excel(path)
    except Exception as e:
        print(f"[excel_to_list] Error: {e}")
        return []
    return df.values.tolist()

def excel_to_dict(file_path: str, sheet_name: str) -> Dict[str, list]:
    '''
    Read a single sheet in an Excel file and return a dictionary 
    with column headers as keys and the rest of the column as the values

    Args:
        file_path (str): Path to the Excel file
        sheet_name (str): Name of the sheet to read from

    Returns:
        dict: Dictionary with column headers as keys and the rest of the column
        as a list of values
    '''
    try:
        df = pd.read_excel(file_path, sheet_name=sheet_name)
        result_dict = {}
        for column in df:
            result_dict[column] = df[column].tolist()
    except Exception as e:
        print(f"[excel_to_dict] Error: {e}")
        return {}
    return result_dict

def excel_to_workbook(file_path: str) -> Union[Workbook, None]:
    '''
    Read an Excel file and return an openpyxl workbook

    Args:
        file_path (str): Path to the Excel file

    Returns:
        openpyxl.workbook.workbook.Workbook: Workbook object
    '''
    try:
        wb = Workbook()
        wb = load_workbook(file_path)
        return wb
    except Exception as e:
        print(f"[excel_to_workbook] Error: {e}")
        return None

def dict_to_excel(input_dict: Dict, save_name: str = 'output.xlsx') -> int:
    '''
    Convert a dictionary to an excel file and save it to the current working directory

    Args:
        input_dict (dict): Dictionary to convert to excel file
        save_name (str): Name of the excel file to save

    Returns:
        int: 0 if successful, 1 if not
    '''
    # Make sure the save name is valid
    if save_name == '':
        print('[dict_to_excel] Error: Save name cannot be empty')
        return 1
    if save_name == 'output.xlsx':
        print('[dict_to_excel] Warning: Save name is the default "output.xlsx". Consider changing it.')
    if not save_name.endswith('.xlsx'):
        save_name += '.xlsx'
        print(f'[dict_to_excel] Warning: Save name does not end with .xlsx. Automatically appending .xlsx to save name')

    # Convert the dictionary to a pandas dataframe
    df = pd.DataFrame(input_dict)

    # Save the dataframe to an excel file
    try:
        df.to_excel(os.path.join(os.getcwd(), save_name), index=False)
        return 0
    except Exception as e:
        print(e)
        return 1
    
def workbook_to_excel(input_workbook: Workbook, save_name: str = 'output.xlsx') -> int:
    '''
    Convert an openpyxl workbook to an excel file and save it to the current working directory
    '''
    
    # Make sure the save name is valid
    if save_name == '':
        print('[openpyxl_to_excel] Error: Save name cannot be empty')
        return 1
    if save_name == 'output.xlsx':
        print('[openpyxl_to_excel] Warning: Save name is the default "output.xlsx". Consider changing it.')
    if not save_name.endswith('.xlsx'):
        save_name += '.xlsx'
        print(f'[openpyxl_to_excel] Warning: Save name does not end with .xlsx. Automatically appending .xlsx to save name')
    
    # Save the workbook to an excel file
    try:
        input_workbook.save(os.path.join(os.getcwd(), save_name))
        return 0
    except Exception as e:
        print(e)
        return 1

def normalize_list(data: Union[List[str], str] = '', keep_originals: bool = False) -> Any:
    '''
    Use a lambda function to normalize the list values by removing all non-alphanumeric characters and converting to lowercase
    Source: https://stackoverflow.com/questions/1276764/stripping-everything-but-alphanumeric-chars-from-a-string-in-python/1276779#1276779

    Args:
        data (Union[List, str]): List of strings or a single string
        keep_originals (bool): If True, return a dictionary with the original values as keys and the normalized values as values
    
    Returns:
        List[str]: List of normalized strings
    '''

    # If the data is a string, convert it to a list
    if isinstance(data, str):
        # Check if the string is empty
        if data == '':
            print('[normalize_list] Warning: Input is an empty string. Returning empty list.')
            return []
        data = [data]

    out_list = list(map(lambda x: ''.join(e for e in x if e.isalnum()).lower(), data))

    # Map the original values to the normalized values
    if keep_originals:
        # map input data : output data in a dictionary
        out_dict = {k: v for k, v in list(zip(data, out_list))}        
    
        return out_list, out_dict
    else:
        return out_list
    
def match_lists(list_1: List, list_2: List, filter_doubles: bool = False) -> Dict:
    '''
    Use difflib to match two lists and returns a nested dictionary with the following structure:
    {
        'list_1_value': {
        'match': 'best_matching_string_in_list_2_here'
        'score': 'ratio_of_similarity_here'
        },
        'another_list_1_value': {
        'match': 'best_matching_string_in_list_2_here'
        'score': 'ratio_of_similarity_here'
        }
    }

    Example output:
    {
        'CO2e emissions': {
        'match': 'CO2e emissons (kg)'
        'score': '0.875'
        },
        'Waste generation': {
        'match': 'Waste genration (kg)'
        'score': '0.973'
        }
    }

    Args:
        list_1 (List): First list of strings
        list_2 (List): Second list of strings
        filter_doubles (bool): If True, make sure each item in list_2 only has one match in list_1

    Returns:
        Dict: Output dictionary with the above structure
    '''

    # Make sure the lists are not empty
    if len(list_1) == 0:
        print('[match_lists] Error: List 1 is empty')
        return {}
    if len(list_2) == 0:
        print('[match_lists] Error: List 2 is empty')
        return {}

    output_dict = {}
    match_relations_dict = {}

    if filter_doubles:
        best_match_strings = []

    for item in list_1:
        best_match_score = 0
        best_match_string = ''

        for item_2 in list_2:
            ratio = difflib.SequenceMatcher(None, item, item_2).ratio()
            if ratio > best_match_score:
                best_match_score = ratio
                best_match_string = item_2

        if filter_doubles:
            if best_match_string in best_match_strings:
                # Check if the new match is better
                print(f"[match_lists] Debug: Matched {item} to {best_match_string} with a score of {np.round(best_match_score, 3)}, but it is already matched to {match_relations_dict[best_match_string]['match']} with a score of {match_relations_dict[best_match_string]['score']}")
                if best_match_score > match_relations_dict[best_match_string]['score']:
                    # Update the output dictionary
                    print(f'[match_lists] Debug: Updating to {item} with a score of {best_match_score}')
                    output_dict[item] = {
                        'match': best_match_string,
                        'score': np.round(best_match_score, 3)
                    }
                    # Update the reverse match for comparison and debugging purposes
                    match_relations_dict[best_match_string] = {
                        'match': item,
                        'score': np.round(best_match_score, 3)
                    }
                else:
                    print(f'[match_lists] Debug: Keeping the previous match with a score of {match_relations_dict[best_match_string]["score"]}')
            else:
                print(f"[match_lists] Debug: Matched {item} to {best_match_string} with a score of {np.round(best_match_score, 3)}")
                # Add the new match to the list of best matches and the output dictionary
                best_match_strings.append(best_match_string)
                output_dict[item] = {
                'match': best_match_string,
                'score': np.round(best_match_score, 3)
                 }
                # Record the reverse match for comparison and debugging purposes
                match_relations_dict[best_match_string] = {
                    'match': item,
                    'score': np.round(best_match_score, 3)
                }
        else:
            output_dict[item] = {
                'match': best_match_string,
                'score': np.round(best_match_score, 3)
            }

    return output_dict