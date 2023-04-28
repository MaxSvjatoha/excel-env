import os
import sys

import numpy as np

from openpyxl import Workbook, load_workbook
import json

import itertools

from typing import List, Dict, Union, Tuple, Any

import difflib


def load_json(json_path: str) -> Dict[str, Any]:
    '''
    Summary:
        Load a JSON file and return its contents as a dictionary.

    Args:
        json_path (str): The path to the JSON file.

    Returns:
        output_json (Dict[str, Any]): A dictionary containing the JSON file contents.

    Raises:
        AssertionError: If the JSON file is not a dictionary.
        IOError: If the JSON file cannot be found or read.
        JSONDecodeError: If the JSON file is not valid JSON.
    '''
    # Load json file:
    with open(json_path, "r") as f:
        output_json = json.load(f)
        assert(type(output_json) == dict)
    
    return output_json


# NOTE: The double underscore (__) prefix indicates that this function is not meant 
# for production use, but this may change in a future update
def __get_input_files(path: Union[List[str], str], settings: Dict) -> List[str]:
    '''
    Summary:
        Get all excel file paths in the given path

    Args:
        path (List[str] or str): Path to the folder(s) to search in
        settings (dict): Script settings dictionary

    Returns:
        list: List of excel files
    '''
    if isinstance(path, str):
        path = [path]
    excel_files = []

    for p in path:
        for root, dirs, files in os.walk(p):
            for f in files:
                # If file is an excel file and is part of a subdirectory, add it to the list
                if f.endswith('.xlsx') and not root.split(os.sep)[-1] in settings['Input file folder name']:
                    excel_files.append(os.path.join(root, f))

    return excel_files


# NOTE: The double underscore (__) prefix indicates that this function is not meant 
# for production use, but this may change in a future update
def __get_input_folders(path: Union[List[str], str], settings: Dict) -> List[str]:
    '''
    Summary:
        Get all folder paths that contain the input excel files

    Args:
        path (List[str] or str): Path to the folder(s) to search in
        settings (dict): Script settings dictionary

    Returns:
        list: List of excel files
    '''
    if isinstance(path, str):
        path = [path]
    excel_folders = []

    for p in path:
        for root, dirs, files in os.walk(p):
            for d in dirs:
                # If dir contains excel files, add it to the list
                if len([f for f in os.listdir(os.path.join(root, d)) if f.endswith('.xlsx')]) > 0:
                    excel_folders.append(os.path.join(root, d))

    return excel_folders


def excel_to_workbook(file_path: str) -> Union[Workbook, None]:
    '''
    Summary:
        Read an Excel file and return an openpyxl workbook

    Args:
        file_path (str): Path to the Excel file

    Returns:
        openpyxl.workbook.workbook.Workbook: Workbook object
    '''
    try:
        wb = Workbook()
        wb = load_workbook(file_path)
        return wb
    except Exception as e:
        print(f"[excel_to_workbook] Error: {e}")
        return None


# Will contain several steps, but for now just removes trailing spaces
def preprocess_cell(cell: str) -> str:
    '''
    Summary:
        Preprocess the given Excel cell using a few hardcoded rules

    Args:
        cell (str): Cell to preprocess

    Returns:
        str: Preprocessed cell
    '''
    return cell.strip()


def match_lists(list_1: List, list_2: List, filter_doubles: bool = False) -> Dict:
    '''
    Use difflib to match two lists and returns a nested dictionary with the following structure:
    {
        'list_1_value': {
        'match': 'best_matching_string_in_list_2_here'
        'score': 'ratio_of_similarity_here'
        },
        'another_list_1_value': {
        'match': 'best_matching_string_in_list_2_here'
        'score': 'ratio_of_similarity_here'
        }
    }

    Example output:
    {
        'CO2e emissions': {
        'match': 'CO2e emissons (kg)'
        'score': '0.875'
        },
        'Waste generation': {
        'match': 'Waste genration (kg)'
        'score': '0.973'
        }
    }

    Args:
        list_1 (List): First list of strings
        list_2 (List): Second list of strings
        filter_doubles (bool): If True, make sure each item in list_2 only has one match in list_1

    Returns:
        Dict: Output dictionary with the above structure
    '''

    # Make sure the lists are not empty
    if len(list_1) == 0:
        print('[match_lists] Error: List 1 is empty')
        return {}
    if len(list_2) == 0:
        print('[match_lists] Error: List 2 is empty')
        return {}

    output_dict = {}
    match_relations_dict = {}

    if filter_doubles:
        best_match_strings = []

    for item in list_1:
        best_match_score = 0
        best_match_string = ''

        for item_2 in list_2:
            ratio = difflib.SequenceMatcher(None, item, item_2).ratio()
            if ratio > best_match_score:
                best_match_score = ratio
                best_match_string = item_2

        if filter_doubles:
            if best_match_string in best_match_strings:
                # Check if the new match is better
                print(f"[match_lists] Debug: Matched {item} to {best_match_string} with a score of {np.round(best_match_score, 3)}, but it is already matched to {match_relations_dict[best_match_string]['match']} with a score of {match_relations_dict[best_match_string]['score']}")
                if best_match_score > match_relations_dict[best_match_string]['score']:
                    # Update the output dictionary
                    print(f'[match_lists] Debug: Updating to {item} with a score of {best_match_score}')
                    output_dict[item] = {
                        'match': best_match_string,
                        'score': np.round(best_match_score, 3)
                    }
                    # Update the reverse match for comparison and debugging purposes
                    match_relations_dict[best_match_string] = {
                        'match': item,
                        'score': np.round(best_match_score, 3)
                    }
                else:
                    print(f'[match_lists] Debug: Keeping the previous match with a score of {match_relations_dict[best_match_string]["score"]}')
            else:
                print(f"[match_lists] Debug: Matched {item} to {best_match_string} with a score of {np.round(best_match_score, 3)}")
                # Add the new match to the list of best matches and the output dictionary
                best_match_strings.append(best_match_string)
                output_dict[item] = {
                'match': best_match_string,
                'score': np.round(best_match_score, 3)
                 }
                # Record the reverse match for comparison and debugging purposes
                match_relations_dict[best_match_string] = {
                    'match': item,
                    'score': np.round(best_match_score, 3)
                }
        else:
            output_dict[item] = {
                'match': best_match_string,
                'score': np.round(best_match_score, 3)
            }

    return output_dict


def get_input_data(input_file_paths: Union[List[str], str], matches: Dict) -> Dict:
    """
    Summary:
        Read the input data from the given Excel files and return a nested dictionary
    Args:
        input_file_paths (Union[List[str], str]): List of paths to the input files
        matches (Dict): Dictionary with the matches between the input and output data
    Returns:
        Dict: Nested dictionary containing the input data
    """
    
    input_data = {}
    for file_path in input_file_paths:
        input_data_key = file_path.split('\\')[-2] # Use the folder name as key
        # Check if key is in the matches dict
        if input_data_key in matches.keys():
            input_data[input_data_key] = {}
        else:
            continue # Immediately go to the next file if key is not in matches

        # Load the workbook at the given path
        wb = excel_to_workbook(file_path)
        
        if wb is None:
            print(f'[get_input_data] Warning: Could not open {file_path}')
            continue

        scope_sheets = [sheet for sheet in wb.sheetnames if 'scope' in sheet.lower()]
        
        if len(scope_sheets) == 0:
            print(f'[get_input_data] Warning: Could not find scope sheet in {file_path}')
            continue
        else:
            for sheet in scope_sheets:
                input_data[input_data_key][sheet] = _get_scope_data(wb, sheet)

        wb.close()
    
    return input_data


# The underscore (_) prefix means that this function is private and is
# only used by modules in this package
# NOTE: Could use some refactoring
def _get_scope_data(wb: Workbook, sheet: str) -> Dict[str, list]:
    '''
    Get the scope data from the given workbook and sheet

    Find key-value pairs by looking and previous and next cells

    If:
    x = colored cell
    o = non-colored cell

    And:
    'o x o' means that: 
    1. Previous cell is not colored
    2. Current cell is colored
    3. Next cell is not colored

    Then:
    Algorithm it triggered by current cell being colored, so middle is always 'x'
    o x o -> key is in the previous cell, value is in the current cell
    o x x -> key is in the next cell, value is in the current cell
    x x o -> skip, since previous cell likely contains a value to some other key
    x x x -> skip, since previous cell likely contains a value to some other key

    Args:
        wb (Workbook): Workbook to get the scope data from
        sheet (str): Worksheet to get the scope data from

    Returns:
        result_dict (Dict[str, list]): Dictionary with the scope data
    '''
    result_dict = {}
    sheet = wb[sheet]
    max_row = sheet.max_row
    max_col = sheet.max_column
    
    for row in range(1, max_row + 1):
        # Don't load the first column, since previous cell can't load then
        for col in range(2, max_col + 1):
            prev_cell = sheet.cell(row=row, column=col-1)
            cell = sheet.cell(row=row, column=col)
            next_cell = sheet.cell(row=row, column=col+1)

            # skip if previous cell is colored (since it likely contains a value to some other key)
            if 'FFDDEBF7'.lower() in prev_cell.fill.start_color.index.lower():
                    continue
            # Check if the current cell color is FFDDEBF7
            elif 'FFDDEBF7'.lower() in cell.fill.start_color.index.lower():
                if next_cell.value is not None:
                    # If it has the same color, then it's the key, else the previous cell is the key
                    if 'FFDDEBF7'.lower() in next_cell.fill.start_color.index.lower():
                        key = next_cell.value
                        #print("[get_scope_data] next cell is key (1)")
                    else:
                        key = prev_cell.value
                        #print("[get_scope_data] prev cell is key (1)")
                # If previous cell is not None, then it's a key, else ignore current cell
                elif prev_cell.value is not None:
                    key = prev_cell.value
                    #print("[get_scope_data] prev cell is key (2)")
                else:
                    continue

                print(f"[get_scope_data] Key: {key}, Value: {cell.value}")
                # Extra step. Possibly temporary until better matching technique are explored:
                # If key ends with " ", remove it
                if key.endswith(' '):
                    key = key[:-1]
                result_dict[key] = cell.value
            else:
                pass # equivalent to 'continue' in this case because end of loop
    return result_dict


# NOTE: Could use some refactoring
def write_data_to_summary(data_dict: Dict, wb: Workbook, matches: Dict, settings: Dict) -> Workbook:
    """
    Writes data from a dictionary to a summary workbook, using a matching dictionary.

    Args:
        data_dict (dict): A dictionary containing the data to be written to the summary workbook.
        wb (openpyxl.Workbook): The summary workbook to write to.
        matches (dict): A dictionary matching keys in data_dict to sheet names in wb.
        settings (dict): A dictionary containing settings for data processing and output.

    Returns:
        openpyxl.Workbook: The modified summary workbook.
    """

    summary_mismatches = wb['Mismatched Data'] # Load the mismatches sheet
    mismatch_count = 0 # Keep track of how many mismatches are found
    mismatch_dict = {} # Nested dict to store keys, items and subitems of mismatches

    for key in data_dict.keys():
        print(f"[write_data_to_summary] Writing data from {key} to sheet {matches[key]['match']}")

        # load sheet from summary file which matches matches[key]['match']
        summary_sheet = wb[matches[key]['match']]
        max_row = summary_sheet.max_row
        max_col = summary_sheet.max_column

        match_count = 0 # Keep track of how many matches are found for a given subitem
        match_dict = {} # Keep track of which row and column a given subitem is found in
        write_key = None # Keep track of which match_dict entry to use when writing data
        write_data = None # Keep track of which data to write to the summary sheet
        special_case = None # Reset special_case

        for item in data_dict[key].keys():
            for subitem in data_dict[key][item].keys():
                for row, col in itertools.product(range(1, max_row + 1), range(1, max_col + 1)):

                    # Load cell value
                    cell_value = summary_sheet.cell(row = row, column = col).value

                    # Pre-process cell value
                    if type(cell_value) == str:
                        cell_value = _preprocess_cell(cell_value)

                    # Try to match subitem to cell_value
                    if subitem == cell_value:
                        #print(f"[write_data_to_summary] Found match for {subitem} at row {row} and column {col}")
                        match_count += 1
                        match_dict[match_count] = {"row": row, "col": col}
                    elif str(subitem) in str(cell_value):
                        #print(f"[write_data_to_summary] {subitem} part of {cell_value} at row {row} and column {col}")
                        match_count += 1
                        match_dict[match_count] = {"row": row, "col": col}
                    else:
                        #print(f"[write_data_to_summary] No match for {subitem} at row {row} and column {col}")
                        pass

                # Handle various amounts of matches
                if match_count > 1:
                    if 'Scope 1'.lower() in item.lower():
                        write_key = min(match_dict.keys()) # use match_dict entry with the lowest key
                    elif 'Scope 3'.lower() in item.lower():
                        write_key = max(match_dict.keys()) # use match_dict entry with the highest key
                    else:
                        print(f"[write_data_to_summary] Multiple matches found for {subitem} in sheet {matches[key]['match']}") 
                        pass # TODO # Handler (non-urgent)
                elif match_count == 0:

                    # Check if subitem is a special case
                    special_case, match_count, write_key, match_dict = _check_if_special_case(item = subitem, match_count = match_count, match_dict = match_dict)
                    
                    if special_case:
                        print(f"[write_data_to_summary] Special case: {special_case}")
                    else:
                        # Register mismatch
                        mismatch_count += 1
                        if key not in mismatch_dict.keys():
                            mismatch_dict[key] = {}
                        if item not in mismatch_dict[key].keys():
                            mismatch_dict[key][item] = {}
                        if subitem not in mismatch_dict[key][item].keys():
                            mismatch_dict[key][item][subitem] = mismatch_count
                            # write to mismatch sheet
                            summary_mismatches.cell(row = mismatch_count+1, column = 1).value = key
                            summary_mismatches.cell(row = mismatch_count+1, column = 2).value = item
                            summary_mismatches.cell(row = mismatch_count+1, column = 3).value = subitem
                            summary_mismatches.cell(row = mismatch_count+1, column = 4).value = "No match found"
                        else:
                            print(f"[write_data_to_summary] {subitem} already in mismatch_dict")
                            pass # TODO # Handler? (this bit was never reached in testing)
                        continue
                else:
                    write_key = list(match_dict.keys())[0] # use match_dict entry with the only key

                # Get data from dict, or generate if missing and settings allow
                if data_dict[key][item][subitem] is not None:
                    write_data = data_dict[key][item][subitem]
                elif settings["Generate missing write data"]:
                    write_data = "GENERATED: " + str(np.random.randint(0, 100))
                else:
                    write_data = None

                # Write data to summary sheet if write_key is not None
                if write_key is not None:
                    summary_sheet.cell(row = match_dict[write_key]["row"], column = match_dict[write_key]["col"] + 1).value = write_data
                    print(f"[write_data_to_summary] Writing {write_data} to row {match_dict[write_key]['row']} and column {match_dict[write_key]['col'] + 1} (sheet: {matches[key]['match']}, {item}, cell name: {subitem})")
                    
                    # Existing value workaround attempt
                    # NOTE: this is the final major issue to be resolved
                    '''
                    # check if cell contains data
                    if summary_sheet.cell(row = match_dict[write_key]["row"], column = match_dict[write_key]["col"] + 1).value is not None:
                        print(f"Row {match_dict[write_key]['row']} and column {match_dict[write_key]['col'] + 1} already contains data: {cell_value}")
                        if type(cell_value) == str:
                            
                            print(f'String trigger for: {summary_sheet.cell(row = match_dict[write_key]["row"], column = match_dict[write_key]["col"] + 1).value}')
                        elif type(cell_value) == int or type(cell_value) == float:

                            print(f'Int/float trigger for: {summary_sheet.cell(row = match_dict[write_key]["row"], column = match_dict[write_key]["col"] + 1).value}')
                    else:

                        print(f"Cell data is None. Writing {write_data} to row {match_dict[write_key]['row']} and column {match_dict[write_key]['col'] + 1}")
                    '''
                else:
                    print(f"[write_data_to_summary] Write key is None for {subitem} in sheet {matches[key]['match']}")

                # Reset match_count and match_dict
                match_count = 0
                match_dict = {}
          
    wb.save(os.path.join(settings['Output file folder path'], settings["Output file name"]))
    wb.close()

    return 0 # Status code 0 if successful
    

# Will contain several steps, but for now just removes trailing spaces
# The underscore (_) prefix means that this function is private and is
# only used by modules in this package
def _preprocess_cell(cell: str) -> str:
    '''
    Preprocess the given cell

    Args:
        cell (str): Cell to preprocess

    Returns:
        str: Preprocessed cell
    '''
    return cell.strip()


# The underscore (_) prefix means that this function is private and is
# only used by modules in this package
def _check_if_special_case(item: str, match_dict: Dict, match_count: int) -> Tuple[Dict, int, str, Dict]:
    """
    Check if the given item matches any of the special cases defined in scope_2_dict.json.

    Args:
        item (str): The item to check.
        match_dict (Dict): A dictionary containing the matches found so far.
        match_count (int): The number of matches found so far.

    Returns:
        A tuple containing the following four elements:
        - special_case (Dict or None): The special case that was matched, or None if no match was found.
        - match_count (int): The updated number of matches found.
        - write_key (str or None): The key to use when retrieving the correct data from the match_dict. 
        None if no match was found.
        - match_dict (Dict): The updated dictionary of matches found.
    """
    special_case = None
    write_key = None
    
    # Load scope 2 special case dictionary (scope_2_dict.json)
    with open('scope_2_dict.json') as f:
        scope_2_dict = json.load(f)

    # Electricity
    if all(x in item.lower() for x in ['källa', 'inköpt el']):
        print(f"[_check_if_special_case] Found 'källa' and 'inköpt el' in subitem: {item}. Using special case 1")
        special_case = scope_2_dict["1"]
        # Note that the write_key is not an int in this case
        match_dict[special_case["name"]] = {"row": special_case["row"], "col": special_case["col"]}
        write_key = special_case["name"]
        match_count += 1
        print(f"{match_count} matches found for {item}")
    elif all(x in item.lower() for x in ['kwh', 'elanvändning']):
        print(f"[_check_if_special_case]  Found 'kwh' and 'elanvändning' in subitem: {item}, using special case 2")
        special_case = scope_2_dict["2"]
        match_dict[special_case["name"]] = {"row": special_case["row"], "col": special_case["col"]}
        write_key = special_case["name"]
        match_count += 1
    # Heat
    elif all(x in item.lower() for x in ['källa', 'värme']):
        print(f"[_check_if_special_case] Found 'källa' and 'värme' in subitem: {item}, using special case 3")
        special_case = scope_2_dict["3"]
        match_dict[special_case["name"]] = {"row": special_case["row"], "col": special_case["col"]}
        write_key = special_case["name"]
        match_count += 1
    elif all(x in item.lower() for x in ['kwh', 'värme']):
        print(f"f[_check_if_special_case] Found 'kwh' and 'värme' in subitem: {item}, using special case 4")
        special_case = scope_2_dict["4"]
        match_dict[special_case["name"]] = {"row": special_case["row"], "col": special_case["col"]}
        write_key = special_case["name"]
        match_count += 1
    # Cooling
    elif all(x in item.lower() for x in ['källa', 'kyla']):
        print(f"[_check_if_special_case] Found 'källa' and 'kyla' in subitem: {item}, using special case 5")
        special_case = scope_2_dict["5"]
        match_dict[special_case["name"]] = {"row": special_case["row"], "col": special_case["col"]}
        write_key = special_case["name"]
        match_count += 1
    elif all(x in item.lower() for x in ['kwh', 'kyla']):
        print(f"[_check_if_special_case] Found 'kwh' and 'kyla' in subitem: {item}, using special case 6")
        special_case = scope_2_dict["6"]
        match_dict[special_case["name"]] = {"row": special_case["row"], "col": special_case["col"]}
        write_key = special_case["name"]
        match_count += 1

    return special_case, match_count, write_key, match_dict