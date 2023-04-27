import os
import sys

import numpy as np
import openpyxl

import itertools

import utils.util as utils

# Setup - script settings:
settings = utils.load_json(json_path="settings.json")

# Add some additional parameters to settings:
settings["Current working directory"] = os.getcwd()
settings["Parent directory"] = os.path.dirname(settings["Current working directory"])
settings["Input file folder path"] = os.path.join(settings["Parent directory"], settings["Input file folder name"])
settings["Output file folder path"] = os.path.join(settings["Parent directory"], settings["Output file folder name"])

# Setup - scope 2 handling settings:
scope_2_dict = utils.load_json(json_path="scope_2_dict.json")

# Setup - Load relevant folder and file paths and extract their names
# Input files and folders
input_folder_paths = utils.__get_input_folders(path = settings["Input file folder path"], settings = settings)
input_file_paths = utils.__get_input_files(path = settings["Input file folder path"], settings = settings)
input_folder_names = [file.split("\\")[-1] for file in input_folder_paths]
input_file_names = [file.split("\\")[-1] for file in input_file_paths]

# Setup - Load summary file and sheets
summary_file = os.path.join(settings["Output file folder path"], settings["Summary file name"])
summary_wb = utils.excel_to_workbook(summary_file)
summary_sheets = summary_wb.sheetnames
if "Mismatched Data" not in summary_sheets:
    summary_mismatches = summary_wb.create_sheet("Mismatched Data")
    summary_mismatches.cell(row = 1, column = 1).value = "Input folder name"
    summary_mismatches.cell(row = 1, column = 2).value = "Scope"
    summary_mismatches.cell(row = 1, column = 3).value = "Entry name"
    summary_mismatches.cell(row = 1, column = 4).value = "Value"
else:
    pass # TODO - decide whether new sheet should be created or if starting row for mismatches 
         # should be set to max_row + 1 or similar

if __name__ == '__main__':

    # Match input file names to summary file sheet names
    matches = utils.match_lists(input_folder_names, summary_sheets, filter_doubles = True)

    print("\nProcessing input files...\n")

    input_data_dict = utils.get_input_data(input_file_paths, matches)

    print("\nWriting data to summary file...\n")

    utils.write_data_to_summary(data_dict = input_data_dict, wb = summary_wb, matches = matches, settings = settings)

    print("\nData write successful. Exiting script...\n")

    #summary_wb.save(os.path.join(settings['Output file folder path'], settings["Output file name"]))
    '''
    for key in input_data_dict.keys():
        # use 'matches' dict to find where to write the data
        print(f"Writing data from input folder '{key}' to sheet '{matches[key]['match']}'")
    '''
    